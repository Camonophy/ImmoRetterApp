"""
Excel export functionality for Kleinanzeigen Scraper
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Optional, List

import pandas as pd
import openpyxl
import openpyxl.utils.exceptions

from .models import ScrapeResult
from config.settings import Settings


# Excel sheet names are limited to 31 chars.
_MAX_SHEET_LEN = 31


def _safe_sheet_name(name: str) -> str:
    """Truncate and sanitize a string to be a valid Excel sheet name."""
    # Excel disallows [ ] : * ? / \ and a few others in sheet names.
    safe = "".join("_" if c in "[]:*?/\\" else c for c in name)
    if len(safe) > _MAX_SHEET_LEN:
        safe = safe[:_MAX_SHEET_LEN]
    return safe


class ExcelExporter:
    """Export scrape results to Excel."""

    def __init__(self, settings: Settings = None):
        self.settings = settings or Settings()

    # ------------------------------------------------------------------
    # Old listings only
    # ------------------------------------------------------------------
    def export_old_listings(self, result: ScrapeResult) -> Optional[str]:
        """
        Export only listings older than MIN_AGE_DAYS.

        Returns the file path, or None if there were no old listings
        (caller is responsible for deciding whether to fall back to
        exporting everything — we don't do that silently here).
        """
        if not result.listings:
            return None

        old_listings = result.get_old_listings()
        if not old_listings:
            return None

        return self._write_excel(
            listings=old_listings,
            result=result,
            kind="old",
            sheet_label="Old Listings",
            filename_template=self.settings.EXCEL_FILENAME_TEMPLATE,
        )

    # ------------------------------------------------------------------
    # All listings
    # ------------------------------------------------------------------
    def export_all_listings(self, result: ScrapeResult) -> Optional[str]:
        """Export every listing (regardless of age)."""
        if not result.listings:
            return None
        return self._write_excel(
            listings=result.listings,
            result=result,
            kind="all",
            sheet_label="All Listings",
            filename_template="{bundesland}_real_estate_listings_{timestamp}.xlsx",
        )

    # ------------------------------------------------------------------
    # Global accumulator — single fixed xlsx across all Bundesländer
    # ------------------------------------------------------------------
    # The "data" columns of the spreadsheet. Defined as a class attribute
    # so other helpers (e.g. the row reader) can refer to the same list.
    GLOBAL_COLUMNS = [
        "Postleitzahl",
        "Name des Verkäufers",
        "Standort",
        "Aktueller Wert (€)",
        "Datum seit online",
        "Link",
    ]

    def export_global(self, result: ScrapeResult,
                      write_all: bool = False) -> Optional[str]:
        """
        Append this run's listings to the global xlsx.

        Listing identity is the URL: any listing whose URL is already
        in the global file is SKIPPED (not updated, not replaced). New
        URLs are appended. The global file therefore only ever grows.

        Args:
            result: The ScrapeResult from a run.
            write_all: If True, write every listing from the run
                (including <90-day-old ones). If False, write only
                listings whose activation date is older than
                MIN_AGE_DAYS. When there are no old listings and
                ``write_all`` is False, this method returns None.

        Layout (matches the per-Bundesland exporter):
          Row 1       Header
          Row 2..N    Listings (existing + new)
          Row N+1     Empty separator
          Row N+2     Today's date in the LAST column
          Sheet 2     "Summary" — this run's stats
        """
        if write_all:
            listings_to_write = result.listings or []
        else:
            listings_to_write = result.get_old_listings()

        if not listings_to_write:
            return None

        filepath = self.settings.OUTPUT_DIR / self.settings.GLOBAL_FILENAME
        existing_rows, existing_urls = self._read_global_rows(filepath)

        # Build the new rows and skip any whose URL is already in the
        # global file. We do NOT update existing rows even if the new
        # run sees a newer activation date — the user explicitly asked
        # for "skip any already added listing".
        new_rows = []
        skipped_duplicate = 0
        for listing in listings_to_write:
            url = listing.url or ""
            if url in existing_urls:
                skipped_duplicate += 1
                continue
            new_rows.append(self._listing_row(listing))
            existing_urls.add(url)

        all_rows = existing_rows + new_rows

        # Record the dedup stats on the result so the Summary sheet
        # shows what happened.
        result.new_global_rows = len(new_rows)
        result.duplicate_global_rows = skipped_duplicate

        return self._write_global_xlsx(
            filepath=filepath,
            rows=all_rows,
            result=result,
            new_rows=len(new_rows),
            skipped=skipped_duplicate,
        )

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------
    def _write_excel(self,
                     listings: List,
                     result: ScrapeResult,
                     kind: str,
                     sheet_label: str,
                     filename_template: str) -> Optional[str]:
        """
        Lay out the workbook in the format requested by the user:

          Row 1:        Header row (Postleitzahl, Name des Verkäufers,
                        Standort, Aktueller Wert (€), Datum seit
                        online, Link)
          Row 2..N+1:   One row per listing with the columns above.
          Row N+2:      Empty separator row (all cells blank).
          Row N+3:      Today's date in the LAST column (Datum seit
                        online) — visually marks the end of the table.
          Summary sheet: unchanged.
        """
        # The order in which we want the columns to appear in the
        # finished Excel file.
        columns = [
            "Postleitzahl",
            "Name des Verkäufers",
            "Standort",
            "Aktueller Wert (€)",
            "Datum seit online",
            "Link",
        ]

        rows = []
        for listing in listings:
            rows.append(self._listing_row(listing))

        df = pd.DataFrame(rows, columns=self.GLOBAL_COLUMNS)

        # Append the empty separator row + the trailing "today" marker.
        # We construct the marker row so that the date string lands in
        # the LAST column (Datum seit online).
        today_str = datetime.now().strftime("%Y-%m-%d")
        empty_row = {col: "" for col in self.GLOBAL_COLUMNS}
        marker_row = dict(empty_row)
        marker_row["Datum seit online"] = today_str
        df = pd.concat(
            [df, pd.DataFrame([empty_row, marker_row],
                               columns=self.GLOBAL_COLUMNS)],
            ignore_index=True,
        )

        timestamp = datetime.now().strftime(self.settings.EXCEL_DATE_FORMAT)
        filename = filename_template.format(
            bundesland=result.bundesland.replace(" ", "_"),
            timestamp=timestamp,
        )
        os.makedirs(self.settings.OUTPUT_DIR, exist_ok=True)
        filepath = self.settings.OUTPUT_DIR / filename

        sheet_name = _safe_sheet_name(f"{result.bundesland} {sheet_label}")

        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                summary_df = pd.DataFrame([result.to_summary_dict()])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
            return str(filepath)
        except (OSError, ValueError,
                openpyxl.utils.exceptions.IllegalCharacterError) as e:
            print(f"Error exporting to Excel: {e}")
            return None

    # ------------------------------------------------------------------
    # Internal helpers (global accumulator)
    # ------------------------------------------------------------------
    def _listing_row(self, listing) -> dict:
        """Turn one Listing into the spreadsheet row dict.

        Single source of truth for row construction — used by both
        the per-Bundesland exporter and the global exporter so the
        column order, the comma price format and the empty-cell
        convention stay in lockstep.
        """
        # Format the price with a thousands separator: 25000 -> "25,000"
        # (per user specification, the separator is a comma, not the
        # German-locale dot). Done on the integer value so we never
        # round-trip a string. When the listing has no parsable price
        # we write an empty cell rather than "0" — empty is honest, 0
        # would look like a real price.
        if listing.price_eur is not None:
            price_cell = f"{int(listing.price_eur):,}"
        else:
            price_cell = ""
        return {
            "Postleitzahl":       listing.postleitzahl or "",
            "Name des Verkäufers":listing.seller_name or "",
            "Standort":           listing.location or "",
            "Aktueller Wert (€)": price_cell,
            "Datum seit online":  listing.date_posted or "",
            "Link":               listing.url or "",
        }

    def _read_global_rows(self, filepath: Path) -> tuple[List[dict], set]:
        """Read the existing global xlsx and return (rows, urls).

        Returns ([], set()) when the file does not exist yet (first run).

        Strips:
          * the header row
          * the trailing empty separator row
          * the trailing "today" marker row (so it does not get treated
            as a duplicate data row on the next run)
        """
        if not filepath.exists():
            return [], set()

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            # Corrupt or unreadable file — start fresh rather than
            # silently dropping user data. Log and continue with an
            # empty set so the next write overwrites cleanly.
            print(f"Warning: could not read global xlsx ({e}); "
                  f"starting fresh.")
            return [], set()

        sheet_name = self.settings.GLOBAL_SHEET_NAME
        if sheet_name not in wb.sheetnames:
            # Older file with a different sheet name — fall back to the
            # first sheet that isn't "Summary".
            data_sheets = [n for n in wb.sheetnames if n != "Summary"]
            if not data_sheets:
                return [], set()
            sheet_name = data_sheets[0]

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return [], set()

        # Identify the trailing marker row (YYYY-MM-DD in the date
        # column with all other cells empty) and the empty separator
        # row above it. Drop both.
        if (rows[-1][4] not in (None, "")
                and "-" in str(rows[-1][4])
                and len(str(rows[-1][4])) == 10
                and all(v in (None, "") for v in (rows[-1][0], rows[-1][1]))):
            rows.pop()
        if rows and all(v is None or v == "" for v in rows[-1]):
            rows.pop()

        idx_link = self.GLOBAL_COLUMNS.index("Link")
        out_rows = []
        urls = set()
        for r in rows:
            # values_only=True gives plain Python values, but defensive
            # str-coercion keeps the set unhashable-safe across openpyxl
            # versions (some return Decimal/Cell for numeric/rich cells).
            row_dict = {
                col: ("" if r[i] is None else str(r[i]))
                for i, col in enumerate(self.GLOBAL_COLUMNS)
            }
            out_rows.append(row_dict)
            urls.add(row_dict["Link"])
        return out_rows, urls

    def _write_global_xlsx(self, filepath: Path, rows: List[dict],
                           result: ScrapeResult,
                           new_rows: int, skipped: int) -> Optional[str]:
        """Write the merged (existing + new) rows to the global xlsx."""
        today_str = datetime.now().strftime("%Y-%m-%d")
        empty_row = {col: "" for col in self.GLOBAL_COLUMNS}
        marker_row = dict(empty_row)
        marker_row["Datum seit online"] = today_str

        all_rows = list(rows) + [empty_row, marker_row]
        df = pd.DataFrame(all_rows, columns=self.GLOBAL_COLUMNS)

        # Extend the summary dict with the dedup stats so the user can
        # see how many listings were new vs. already-known.
        summary = dict(result.to_summary_dict())
        summary["New Global Rows"] = new_rows
        summary["Duplicate (skipped)"] = skipped
        summary_df = pd.DataFrame([summary])

        sheet_name = self.settings.GLOBAL_SHEET_NAME

        os.makedirs(self.settings.OUTPUT_DIR, exist_ok=True)
        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
            return str(filepath)
        except (OSError, ValueError,
                openpyxl.utils.exceptions.IllegalCharacterError) as e:
            print(f"Error exporting global xlsx: {e}")
            return None


# ----------------------------------------------------------------------
# Module-level helper
# ----------------------------------------------------------------------
def export_to_excel(result: ScrapeResult, allow_all_fallback: bool = False) -> Optional[str]:
    """
    Convenience export function.

    By default this routes to :py:meth:`ExcelExporter.export_global`,
    which writes the run's listings into the single global xlsx
    (data/output/Global_real_estate_old_listings.xlsx) and skips any
    listing whose URL is already present. The file accumulates across
    runs and across Bundesländer.

    For backward compatibility the old per-Bundesland writers are
    still available as :py:meth:`ExcelExporter.export_old_listings`
    and :py:meth:`ExcelExporter.export_all_listings` — instantiate
    ``ExcelExporter()`` directly if you need them.

    ``allow_all_fallback`` is honoured the same way it was before:
    if True (i.e. ``--all`` on the CLI), the global file receives
    *every* listing from the run, not just the >90-day ones.
    """
    exporter = ExcelExporter()
    return exporter.export_global(result, write_all=allow_all_fallback)